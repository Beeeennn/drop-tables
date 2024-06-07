import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Superstore Dataset.xlsx")
# Define variable to read sheet
dataframe1 = dataframe.active

# declare arrays to contain values
Catergories = []
SubCatergories = []
Profits = []
SubProfits = []
AvrageProfits = []
SubAvrageProfits = []
Sales = []
SubSales = []

# Iterate the loop to read the cell values
for Row in range(2, dataframe1.max_row):

	# cheack that the row contains information
	if dataframe1.cell(row=Row, column=1).value is not None:

		# defining whitch cells to look at
		CategoryCell = dataframe1.cell(row=Row, column=15)
		SubCategoryCell = dataframe1.cell(row=Row, column=16)
		ProfitCell = dataframe1.cell(row=Row, column=21)

		# extending lists if category is not already included
		if CategoryCell.value not in Catergories:
			Catergories.append(CategoryCell.value)
			Profits.append(0)
			Sales.append(0)
			AvrageProfits.append(0)

		# repeat for sub categories
		if SubCategoryCell.value not in SubCatergories:
			SubCatergories.append(SubCategoryCell.value)
			SubProfits.append(0)
			SubSales.append(0)
			SubAvrageProfits.append(0)

		# counting profits and sales per category
		Profits[Catergories.index(CategoryCell.value)] += ProfitCell.value
		Sales[Catergories.index(CategoryCell.value)] += dataframe1.cell(row=Row, column=19).value

		# repeat for sub categories
		SubProfits[SubCatergories.index(SubCategoryCell.value)] += ProfitCell.value
		SubSales[SubCatergories.index(SubCategoryCell.value)] += dataframe1.cell(row=Row, column=19).value

# calculating average profit per item sold in each category
for i in range (0, len(Profits)):
	AvrageProfits[i] = Profits[i] / Sales[i]

# print result
print("Catergories")
print(Catergories)
print("Total Profit")
print(Profits)
print("average profit per item")
print(AvrageProfits)

print("\n")

# print highest profit and profit per item
print ("the category with the highest total profit is", Catergories[Profits.index(max(Profits))], "with" , Profits[Profits.index(max(Profits))], "total profit made")
print ("the category with the highest average profit is", Catergories[AvrageProfits.index(max(AvrageProfits))], "with" , AvrageProfits[AvrageProfits.index(max(AvrageProfits))], "average profit per usit sold")


# repeat for subcategories
print("\n")

for i in range (0, len(SubProfits)):
	SubAvrageProfits[i] = SubProfits[i] / SubSales[i]

print("Sub Catergories")
print(SubCatergories)
print("Total Profit")
print(SubProfits)
print("average profit per item")
print(SubAvrageProfits)

print("\n")

print ("the sub catergory with the highest total profit is", SubCatergories[SubProfits.index(max(SubProfits))], "with" , SubProfits[SubProfits.index(max(SubProfits))], "total profit made")
print ("the sub catergory with the highest average profit is", SubCatergories[SubAvrageProfits.index(max(SubAvrageProfits))], "with" , SubAvrageProfits[SubAvrageProfits.index(max(SubAvrageProfits))], "average profit per usit sold")